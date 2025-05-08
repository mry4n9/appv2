import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import streamlit as st
import json

def apply_header_style(cell):
    """Applies header style to a cell."""
    cell.font = Font(color="FFFFFF", bold=True)
    cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

def apply_default_cell_style(cell, is_version_col=False):
    """Applies default style to a data cell."""
    if is_version_col:
        cell.alignment = Alignment(horizontal="center", vertical="center")
    else:
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True) # Vertical align middle, expand cells
    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

def set_column_widths_and_row_heights(sheet):
    """Adjusts column widths and row heights for readability."""
    for col_idx, column_cells in enumerate(sheet.columns):
        max_length = 0
        column_letter = get_column_letter(col_idx + 1)
        
        # Check header length
        header_cell = sheet[f"{column_letter}1"]
        if header_cell.value:
            header_text_lines = str(header_cell.value).split('\n')
            max_header_line_len = max(len(line) for line in header_text_lines) if header_text_lines else 0
            max_length = max(max_length, max_header_line_len)

        for cell in column_cells:
            if cell.row == 1: continue # Skip header for content length check
            try:
                if cell.value:
                    cell_text_lines = str(cell.value).split('\n')
                    max_cell_line_len = max(len(line) for line in cell_text_lines) if cell_text_lines else 0
                    if max_cell_line_len > max_length:
                        max_length = max_cell_line_len
            except:
                pass
        
        adjusted_width = (max_length + 5) * 1.2 # Add padding
        if column_letter == 'A': # Version #
             adjusted_width = 10
        sheet.column_dimensions[column_letter].width = min(adjusted_width, 70) # Max width 70

    for row_idx, row_cells in enumerate(sheet.rows):
        max_lines = 1
        for cell in row_cells:
            if cell.value:
                lines = len(str(cell.value).split('\n'))
                if lines > max_lines:
                    max_lines = lines
        # Approximate row height: 15 points per line, plus some padding
        sheet.row_dimensions[row_idx + 1].height = max(20, max_lines * 15 + 5)


def create_excel_file(content_data: dict, company_name: str, lead_objective: str, company_info_for_reasoning: dict) -> bytes:
    """Creates an Excel file with structured ad content."""
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Email Page
    if "email" in content_data and content_data["email"]:
        ws_email = wb.create_sheet("Email")
        headers_email = ["Version #", "Objective", "Headline", "Subject Line", "Body", "CTA"]
        for col_num, header in enumerate(headers_email, 1):
            cell = ws_email.cell(row=1, column=col_num, value=header)
            apply_header_style(cell)

        for row_num, data_row in enumerate(content_data["email"], 2):
            if isinstance(data_row, dict) and "error" not in data_row:
                ws_email.cell(row=row_num, column=1, value=data_row.get("version_number")).alignment = Alignment(horizontal="center", vertical="center")
                ws_email.cell(row=row_num, column=2, value=data_row.get("objective"))
                ws_email.cell(row=row_num, column=3, value=data_row.get("headline"))
                ws_email.cell(row=row_num, column=4, value=data_row.get("subject_line"))
                ws_email.cell(row=row_num, column=5, value=data_row.get("body"))
                ws_email.cell(row=row_num, column=6, value=data_row.get("cta"))
            else: # Handle error placeholder
                ws_email.cell(row=row_num, column=1, value="Error").alignment = Alignment(horizontal="center", vertical="center")
                for i in range(2, len(headers_email) + 1):
                    ws_email.cell(row=row_num, column=i, value=str(data_row))


        for row in ws_email.iter_rows(min_row=1, max_row=ws_email.max_row, min_col=1, max_col=len(headers_email)):
            for idx, cell in enumerate(row):
                apply_default_cell_style(cell, is_version_col=(idx == 0 and cell.row > 1))
        set_column_widths_and_row_heights(ws_email)

    # LinkedIn Page
    if "linkedin" in content_data and content_data["linkedin"]:
        ws_linkedin = wb.create_sheet("LinkedIn")
        headers_linkedin = ["Version #", "Ad Name", "Objective", "Introductory Text", "Image Copy", "Headline", "Destination", "CTA Button"]
        for col_num, header in enumerate(headers_linkedin, 1):
            cell = ws_linkedin.cell(row=1, column=col_num, value=header)
            apply_header_style(cell)
        
        for row_num, data_row in enumerate(content_data["linkedin"], 2):
            if isinstance(data_row, dict) and "error" not in data_row:
                ws_linkedin.cell(row=row_num, column=1, value=data_row.get("version_number")).alignment = Alignment(horizontal="center", vertical="center")
                ws_linkedin.cell(row=row_num, column=2, value=data_row.get("ad_name"))
                ws_linkedin.cell(row=row_num, column=3, value=data_row.get("objective"))
                ws_linkedin.cell(row=row_num, column=4, value=data_row.get("introductory_text"))
                ws_linkedin.cell(row=row_num, column=5, value=data_row.get("image_copy"))
                ws_linkedin.cell(row=row_num, column=6, value=data_row.get("headline"))
                ws_linkedin.cell(row=row_num, column=7, value=data_row.get("destination"))
                ws_linkedin.cell(row=row_num, column=8, value=data_row.get("cta_button"))
            else:
                ws_linkedin.cell(row=row_num, column=1, value="Error").alignment = Alignment(horizontal="center", vertical="center")
                for i in range(2, len(headers_linkedin) + 1):
                    ws_linkedin.cell(row=row_num, column=i, value=str(data_row))


        for row in ws_linkedin.iter_rows(min_row=1, max_row=ws_linkedin.max_row, min_col=1, max_col=len(headers_linkedin)):
            for idx, cell in enumerate(row):
                apply_default_cell_style(cell, is_version_col=(idx == 0 and cell.row > 1))
        set_column_widths_and_row_heights(ws_linkedin)

    # Facebook Page
    if "facebook" in content_data and content_data["facebook"]:
        ws_facebook = wb.create_sheet("Facebook")
        headers_facebook = ["Version #", "Ad Name", "Objective", "Primary Text", "Image Copy", "Headline", "Link Description", "Destination", "CTA Button"]
        for col_num, header in enumerate(headers_facebook, 1):
            cell = ws_facebook.cell(row=1, column=col_num, value=header)
            apply_header_style(cell)

        for row_num, data_row in enumerate(content_data["facebook"], 2):
            if isinstance(data_row, dict) and "error" not in data_row:
                ws_facebook.cell(row=row_num, column=1, value=data_row.get("version_number")).alignment = Alignment(horizontal="center", vertical="center")
                ws_facebook.cell(row=row_num, column=2, value=data_row.get("ad_name"))
                ws_facebook.cell(row=row_num, column=3, value=data_row.get("objective"))
                ws_facebook.cell(row=row_num, column=4, value=data_row.get("primary_text"))
                ws_facebook.cell(row=row_num, column=5, value=data_row.get("image_copy"))
                ws_facebook.cell(row=row_num, column=6, value=data_row.get("headline"))
                ws_facebook.cell(row=row_num, column=7, value=data_row.get("link_description"))
                ws_facebook.cell(row=row_num, column=8, value=data_row.get("destination"))
                ws_facebook.cell(row=row_num, column=9, value=data_row.get("cta_button"))
            else:
                ws_facebook.cell(row=row_num, column=1, value="Error").alignment = Alignment(horizontal="center", vertical="center")
                for i in range(2, len(headers_facebook) + 1):
                    ws_facebook.cell(row=row_num, column=i, value=str(data_row))

        for row in ws_facebook.iter_rows(min_row=1, max_row=ws_facebook.max_row, min_col=1, max_col=len(headers_facebook)):
            for idx, cell in enumerate(row):
                apply_default_cell_style(cell, is_version_col=(idx == 0 and cell.row > 1))
        set_column_widths_and_row_heights(ws_facebook)

    # Google Search Page
    if "google_search" in content_data and content_data["google_search"]:
        gs_data = content_data["google_search"][0] if content_data["google_search"] else {} # Expects a list with one item
        if isinstance(gs_data, dict) and "error" not in gs_data:
            ws_gsearch = wb.create_sheet("Google Search")
            headers_gsearch = ["Headline", "Description"]
            for col_num, header in enumerate(headers_gsearch, 1):
                cell = ws_gsearch.cell(row=1, column=col_num, value=header)
                apply_header_style(cell)

            headlines = gs_data.get("headlines", [])
            descriptions = gs_data.get("descriptions", [])
            max_rows = max(len(headlines), len(descriptions))

            for i in range(max_rows):
                if i < len(headlines):
                    ws_gsearch.cell(row=i + 2, column=1, value=headlines[i])
                if i < len(descriptions):
                    ws_gsearch.cell(row=i + 2, column=2, value=descriptions[i])
                else: # Grey out empty description cells
                    cell = ws_gsearch.cell(row=i + 2, column=2)
                    cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid") # Light grey

            for row in ws_gsearch.iter_rows(min_row=1, max_row=ws_gsearch.max_row, min_col=1, max_col=len(headers_gsearch)):
                for cell in row: apply_default_cell_style(cell)
            set_column_widths_and_row_heights(ws_gsearch)

    # Google Display Page
    if "google_display" in content_data and content_data["google_display"]:
        gd_data = content_data["google_display"][0] if content_data["google_display"] else {} # Expects a list with one item
        if isinstance(gd_data, dict) and "error" not in gd_data:
            ws_gdisplay = wb.create_sheet("Google Display")
            headers_gdisplay = ["Headline", "Description"]
            for col_num, header in enumerate(headers_gdisplay, 1):
                cell = ws_gdisplay.cell(row=1, column=col_num, value=header)
                apply_header_style(cell)

            headlines = gd_data.get("headlines", [])
            descriptions = gd_data.get("descriptions", [])
            max_rows = max(len(headlines), len(descriptions)) # Should be 5 for both

            for i in range(max_rows):
                if i < len(headlines): ws_gdisplay.cell(row=i + 2, column=1, value=headlines[i])
                if i < len(descriptions): ws_gdisplay.cell(row=i + 2, column=2, value=descriptions[i])
            
            for row in ws_gdisplay.iter_rows(min_row=1, max_row=ws_gdisplay.max_row, min_col=1, max_col=len(headers_gdisplay)):
                for cell in row: apply_default_cell_style(cell)
            set_column_widths_and_row_heights(ws_gdisplay)

    # Reasoning Page
    ws_reasoning = wb.create_sheet("Reasoning")
    current_row = 1
    
    # Header for Scraped Data section
    header_scraped_cell = ws_reasoning.cell(row=current_row, column=1, value="Scraped Company Information")
    apply_header_style(header_scraped_cell)
    ws_reasoning.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
    current_row += 1

    for key, value in company_info_for_reasoning.items():
        key_cell = ws_reasoning.cell(row=current_row, column=1, value=key.replace("_", " ").title())
        key_cell.font = Font(bold=True)
        apply_default_cell_style(key_cell)
        
        val_str = ""
        if isinstance(value, list):
            val_str = "\n".join(f"- {item}" for item in value) if value else "N/A"
        else:
            val_str = str(value) if value else "N/A"
        
        value_cell = ws_reasoning.cell(row=current_row, column=2, value=val_str)
        apply_default_cell_style(value_cell)
        current_row += 1
    
    current_row += 1 # Add a blank row for spacing

    # Header for AI Reasoning
    header_reasoning_cell = ws_reasoning.cell(row=current_row, column=1, value="AI Generated Reasoning")
    apply_header_style(header_reasoning_cell)
    ws_reasoning.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
    current_row += 1

    reasoning_text = "No reasoning generated."
    if "reasoning" in content_data and content_data["reasoning"]:
        if isinstance(content_data["reasoning"], dict) and "reasoning_statement" in content_data["reasoning"]:
            reasoning_text = content_data["reasoning"]["reasoning_statement"]
        elif isinstance(content_data["reasoning"], str): # Fallback if it's just a string
            reasoning_text = content_data["reasoning"]

    reasoning_cell = ws_reasoning.cell(row=current_row, column=1, value=reasoning_text)
    apply_default_cell_style(reasoning_cell)
    ws_reasoning.merge_cells(start_row=current_row, start_column=1, end_row=current_row + 5, end_column=2) # Merge for more space
    
    set_column_widths_and_row_heights(ws_reasoning)


    # Save to a BytesIO object
    excel_bytes = io.BytesIO()
    wb.save(excel_bytes)
    excel_bytes.seek(0)
    return excel_bytes.getvalue()